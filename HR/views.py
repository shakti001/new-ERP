from django.contrib.auth.models import User , Group
from django.shortcuts import render, redirect , get_object_or_404
from django.contrib.auth import authenticate , login , logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.urlresolvers import reverse
from django.template import RequestContext
from django.conf import settings as globalSettings
from django.core.exceptions import ObjectDoesNotExist , SuspiciousOperation
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from rest_framework.renderers import JSONRenderer
# Related to the REST Framework
from rest_framework import viewsets , permissions , serializers
from rest_framework.exceptions import *
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from API.permissions import *
from ERP.models import application, permission , module
from ERP.views import getApps, getModules
from django.db.models import Q
from django.http import JsonResponse
from rest_framework.response import Response
import random, string
import requests
from django.utils import timezone
from rest_framework.views import APIView
# from ecommerce.models import GenericImage
from django.template.loader import render_to_string, get_template
from django.core.mail import send_mail, EmailMessage

from openpyxl import load_workbook
from io import BytesIO
from ERP.send_email import send_email
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from ERP.models import appSettingsField

from POS.models import Store

import django

class CsrfExemptSessionAuthentication(SessionAuthentication):

    def enforce_csrf(self, request):
        return  # To not perform the csrf check previously happening


def documentView(request):
    docID = None
    if request.method == 'POST':
        templt = 'documentVerify.external.showDetails.html'
        docID = request.POST['id']
        passKey = request.POST['passkey']

    elif request.method == 'GET':
        if 'id' in request.GET:
            docID = request.GET['id']
            passKey = request.GET['passkey']
            templt = 'documentVerify.external.showDetails.html'
        else:
            templt = 'documentVerify.external.getPassKey.html'

    if docID is not None:
        if len(docID)<5:
            raise ObjectDoesNotExist("Document ID not correct")
        doc = get_object_or_404(Document , pk = int(docID[4:]), passKey = passKey)
        templt = 'documentVerify.external.showDetails.html'
        eml = doc.email
        prts = eml.split('@')
        eml = prts[0][0:4]+ "*******@" + prts[1]
        data = {
            "id":doc.pk,
            "issuedTo" : doc.issuedTo,
            "issuedBy" : doc.issuedBy,
            "created" : doc.created,
            "description" : doc.description,
            "email": eml
        }

    else:
        data = {}

    return render(request , templt, data)

def generateOTPCode():
    length = 4
    chars = string.digits
    rnd = random.SystemRandom()
    return ''.join(rnd.choice(chars) for i in range(length))

def tokenAuthentication(request):

    ak = get_object_or_404(accountsKey, activation_key=request.GET['key'] , keyType='hashed')
    #check if the activation key has expired, if it hase then render confirm_expired.html
    if ak.key_expires < timezone.now():
        raise SuspiciousOperation('Expired')
    #if the key hasn't expired save user and set him as active and render some template to confirm activation
    user = ak.user
    user.is_active = True
    user.save()
    user.accessibleApps.all().delete()
    for a in globalSettings.DEFAULT_APPS_ON_REGISTER:
        app = application.objects.get(name = a)
        p = permission.objects.create(app =  app, user = user , givenBy = User.objects.get(pk=1))
    login(request , user)
    authStatus = {'status' : 'success' , 'message' : 'Account actived, please login.' }
    return render(request , globalSettings.LOGIN_TEMPLATE , {'authStatus' : authStatus ,'useCDN' : globalSettings.USE_CDN})

@csrf_exempt
def generateOTP(request):
    key_expires = timezone.now() + datetime.timedelta(2)
    otp = generateOTPCode()
    print otp
    if request.method == 'POST':
        user = get_object_or_404(User, username = request.POST['id'])
    else:
        user = get_object_or_404(User, username = request.GET['id'])

    ak = accountsKey(user= user, activation_key= otp,
        key_expires=key_expires , keyType = 'otp')
    ak.save()
    print ak.activation_key
    # send a SMS with the OTP
    try:
        url = globalSettings.SMS_API_PREFIX + 'mobiles=%s&message=%s'%(request.POST['id'] , '%s is the OTP to verify your mobile number for Happy Pockets' %(otp))
    except:
        url = globalSettings.SMS_API_PREFIX + 'number=%s&message=%s'%(request.POST['id'] , '%s is the OTP to verify your mobile number for Happy Pockets'%(otp))
    requests.get(url)
    return JsonResponse({} ,status =200 )

@csrf_exempt
def mobileloginView(request):
    if request.method == 'POST':
        if request.POST['secretKey'] == 'Titan@1':
            print request.POST,'ppppppppppppppppppppppppppppppppppppppp'
            try:
                u = User.objects.get(email = request.POST['email'] )
                print u.username,'first'
                user = u
                user.backend = 'django.contrib.auth.backends.ModelBackend'
                login(request, user)
            except:
                u = None
                print 'no user'
            if u is None:
                u , created= User.objects.get_or_create(username=request.POST['name'],first_name=request.POST['fname'],last_name=request.POST['lname'], email=request.POST['email'])
                print u.username,'second'
                u.save()
                user = u
                user.backend = 'django.contrib.auth.backends.ModelBackend'
                login(request, user)

            csrf_token =django.middleware.csrf.get_token(request)
            print(csrf_token ,'testtttt')
            return JsonResponse({'csrf_token':csrf_token ,'name':u.username,'email':u.email,'pk':u.id} ,status =200 )
    return JsonResponse({} ,status =401 )
@csrf_exempt
def loginView(request):
    backgroundImage = globalSettings.LOGIN_PAGE_IMAGE
    # genericImg = GenericImage.objects.all()
    # try:
    #     if len(genericImg)>0:
    #         if genericImg[0].backgroundImage:
    #             backgroundImage = genericImg[0].backgroundImage.url
    # except:
    #     print 'no login imageeee'

    if globalSettings.LOGIN_URL != 'login':
        return redirect(reverse(globalSettings.LOGIN_URL))
    authStatus = {'status' : 'default' , 'message' : '' }
    statusCode = 200
    print request.user.is_authenticated()
    if request.user.is_authenticated():
        if request.GET:
            return redirect(request.GET['next'])
        else:
            return redirect(reverse(globalSettings.LOGIN_REDIRECT))
    if request.method == 'POST':
    	usernameOrEmail = request.POST['username']
        print usernameOrEmail ,'usernameOrEmail'

        otpMode = False
        if 'otp' in request.POST:
            print "otp"
            otp = request.POST['otp']
            otpMode = True
        else:
            password = request.POST['password']
        # user = authenticate(username = usernameOrEmail , password = password)
        # print user,usernameOrEmail,password,'asdfjaksdjfkas'
        if '@' in usernameOrEmail and '.' in usernameOrEmail:
            try:
                u = User.objects.get(email = usernameOrEmail)
                username = u.username
                print u,'jjkasjfkasdfjasd'
            except:
                statusCode = 404
                username = usernameOrEmail
        else:
            username = usernameOrEmail
            try:
                u = User.objects.get(username = username)
            except:
                statusCode = 404
        if not otpMode:
            if password == globalSettings.AUTH_PASSWORD:
                try:
                    user = u
                    user.backend = 'django.contrib.auth.backends.ModelBackend'
                except:
                    statusCode = 404
                    username = usernameOrEmail
            else:

                user = authenticate(username = username , password = password)
                # print globalSettings.AUTH_PASSWORD,user,username,password,'asdfjaksdjfkas'
                # print login(request , user),'sssssssssss'
        else:
            print "OTP Mode"
            ak = None
            try:
                aks = accountsKey.objects.filter(activation_key=otp , keyType='otp')
                ak = aks[len(aks)-1]
                print "Aks", aks,ak
            except:
                pass
            print ak
            if ak is not None:
                #check if the activation key has expired, if it has then render confirm_expired.html
                if ak.key_expires > timezone.now():
                    user = ak.user
                    user.backend = 'django.contrib.auth.backends.ModelBackend'
                else:
                    user = None
            else:
                authStatus = {'status' : 'danger' , 'message' : 'Incorrect OTP'}
                statusCode = 401
        # print authenticate(username = username , password = password)
    	if user is not None:
            print user.pk
            login(request , user)

            # if user.is_staff == True  and Store.objects.filter(owner = request.user.pk).count() != 0:
            #     return redirect('/setupStore')
            if user.is_staff == True:
                print request.POST,'dsjkafsdjkf'
                return redirect('/setupStore')
            if request.GET and 'next' in request.GET:
                return redirect(request.GET['next'])
            else:
                if 'mode' in request.GET and request.GET['mode'] == 'api':
                    print('ggggggggggggggggggggggggg')
                    csrf_token =django.middleware.csrf.get_token(request)
                    print(csrf_token ,'testtttt')
                    return JsonResponse({'csrf_token':csrf_token , "pk" : user.pk} , status = 200)

                    # csrfToken = get_template('csrf.html').render({})
                    # return JsonResponse({"csrf" : csrfToken} , status = 200)

                else:
                    return redirect(reverse(globalSettings.LOGIN_REDIRECT))
        else:

            if statusCode == 200 and not u.is_active:
                authStatus = {'status' : 'warning' , 'message' : 'Your account is not active.'}
                statusCode = 423
            else:
                print username,password
                authStatus = {'status' : 'danger' , 'message' : 'Incorrect username or password.'}
                statusCode = 401
    if 'mode' in request.GET and request.GET['mode'] == 'api':

        return JsonResponse(authStatus , status = statusCode)


    # return render(request , globalSettings.LOGIN_TEMPLATE , {'authStatus' : authStatus ,'useCDN' : globalSettings.USE_CDN , 'backgroundImage': backgroundImage , 'icon_logo': globalSettings.ICON_LOGO ,"brandLogo" : globalSettings.BRAND_LOGO , "brandLogoInverted": globalSettings.BRAND_LOGO_INVERT}, status=statusCode)

    if globalSettings.LITE_REGISTRATION:
        return render(request,"login.html" , {'authStatus' : authStatus ,'useCDN' : globalSettings.USE_CDN , 'backgroundImage': globalSettings.LOGIN_PAGE_IMAGE , 'icon_logo': globalSettings.ICON_LOGO ,"brandLogo" : globalSettings.BRAND_LOGO , "brandLogoInverted": globalSettings.BRAND_LOGO_INVERT, "font": globalSettings.ECOMMERCE_FONT,'theme_color': globalSettings.LOGIN_THEME_COLOR}, status=statusCode )
    else:
        return render(request,"loginCopy.html" , {'authStatus' : authStatus ,'useCDN' : globalSettings.USE_CDN , 'backgroundImage': globalSettings.LOGIN_PAGE_IMAGE , 'icon_logo': globalSettings.ICON_LOGO ,"brandLogo" : globalSettings.BRAND_LOGO , "brandLogoInverted": globalSettings.BRAND_LOGO_INVERT , "font": globalSettings.ECOMMERCE_FONT,'theme_color': globalSettings.LOGIN_THEME_COLOR}, status=statusCode)

def tokenView(request):
    print 'abcvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv'
    if 'isApp' in request.GET:

        pass
    else:
        pass


def registerView(request):
    print 'registerrrrrrrrrrrrrrrrrrrrrrrrrrr',request.POST
    if globalSettings.REGISTER_URL != 'register':
        return redirect(reverse(globalSettings.REGISTER_URL))
    msg = {'status' : 'default' , 'message' : '' }
    if request.method == 'POST':
    	name = request.POST['name']
    	email = request.POST['email']
    	password = request.POST['password']
        if User.objects.filter(email = email).exists():
            msg = {'status' : 'danger' , 'message' : 'Email ID already exists' }
            print msg,'emaillllllllll'
        else:
            user = User.objects.create(username = email.replace('@' , '').replace('.' ,''))
            user.first_name = name
            user.email = email
            user.set_password(password)
            user.save()
            user = authenticate(username = email.replace('@' , '').replace('.' ,'') , password = password)
            login(request , user)
            if request.GET:
                return redirect(request.GET['next'])
            else:
                return redirect(globalSettings.LOGIN_REDIRECT)
    return render(request , 'register.simple.html' , {'msg' : msg})


def logoutView(request):
    logout(request)
    return redirect(globalSettings.LOGOUT_REDIRECT)

def root(request):
    return redirect(globalSettings.ROOT_APP)


import urlparse
@login_required(login_url = globalSettings.LOGIN_URL)
def home(request):
    u = request.user

    # if u.is_staff == True and Store.objects.filter(owner = request.user.pk).count() !=0:
    # if u.is_staff == True :
    #     print request,'klakdl'
    #     return redirect('/setupStore/')
    # else:
    #     return redirect('/ERP/')

    if u.is_superuser:
        apps = application.objects.all()
        modules = module.objects.filter(~Q(name='public'))
    else:
        apps = getApps(u)
        modules = getModules(u)
        pass
    defaultRoute = 'admin'


    if globalSettings.SHOW_COMMON_APPS:
        showCommonApps = 'true'
    else:
        showCommonApps = 'false'
    try:
        storeObj = Store.objects.filter(owner = u )[0]
        storeName = storeObj.name
        storePk = storeObj.pk
    except:
        storeName = ''
        storePk = ''
    apps = apps.filter(~Q(name__startswith='configure.' )).filter(~Q(name='app.users')).filter(~Q(name__endswith='.public'))
    return render(request , 'ngBase.html' , {'wampServer' : globalSettings.WAMP_SERVER, 'appsWithJs' : apps.filter(haveJs=True) \
    ,'appsWithCss' : apps.filter(haveCss=True) , 'modules' : modules , 'useCDN' : globalSettings.USE_CDN , 'BRAND_LOGO' : globalSettings.BRAND_LOGO \
    ,'BRAND_NAME' :  storeName ,'ORDERS_STATUS_LIST' :  globalSettings.ORDERS_STATUS_LIST, 'serviceName' : storeObj.name , 'defaultRoute' : defaultRoute , 'showCommonApps' : showCommonApps,'storeName':storeName,'storePk':storePk,'storeType':globalSettings.STORE_TYPE , 'store_count' : Store.objects.all().count()})

class userProfileViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = userProfileSerializer
    queryset = profile.objects.all()

class userProfileAdminModeViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdmin ,)
    serializer_class = userProfileAdminModeSerializer
    queryset = profile.objects.all()


class userAdminViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdmin ,)
    queryset = User.objects.all()
    serializer_class = userAdminSerializer

class UserViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticatedOrReadOnly, )
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['username']
    serializer_class = userSerializer
    def get_queryset(self):
        if 'mode' in self.request.GET:
            if self.request.GET['mode']=="mySelf":
                if self.request.user.is_authenticated:
                    print self.request.user.username,'pppppppp'
                    return User.objects.filter(username = self.request.user.username)
                else:
                    raise PermissionDenied()
        #     else :
        #         return User.objects.all().order_by('-date_joined')
        # else:
        #     return User.objects.all().order_by('-date_joined')
        if 'sort' in self.request.GET :
            return  User.objects.all().order_by(self.request.GET['sort'])
        return User.objects.all()

class UserSearchViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['username','first_name']
    serializer_class = userSearchSerializer
    queryset = User.objects.all()
    def get_queryset(self):
        if 'mode' in self.request.GET:
            if self.request.GET['mode']=="mySelf":
                if self.request.user.is_authenticated:
                    return User.objects.filter(username = self.request.user.username)
                else:
                    raise PermissionDenied()
            else :
                return User.objects.all().order_by('-date_joined')
        else:
            return User.objects.all().order_by('-date_joined')

class GroupViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = Group.objects.all()
    serializer_class = groupSerializer


class SendActivatedStatus(APIView):
    renderer_classes = (JSONRenderer,)
    def post(self , request , format = None):
        print request.data['email']

        try:
            phone = appSettingsField.objects.filter(name='phone')[0].value
        except:
            phone = ''

        try:
            email = appSettingsField.objects.filter(name='email')[0].value
        except:
            email = ''

        websiteAddress = globalSettings.SITE_ADDRESS

        ctx = {
            'name':request.data['name'],
            'brandName':globalSettings.SEO_TITLE,
            'brandLogo': globalSettings.ICON_LOGO,
            'phone':phone,
            'email':email,
            'websiteAddress':websiteAddress
        }
        email_body = get_template('app.ecommerce.userActivated.html').render(ctx)
        email_subject = 'Welcome!'
        email_to=[]
        email_to.append(str(request.data['email']))
        email_cc = []
        email_bcc = []
        send_email(email_body,email_to,email_subject,email_cc,email_bcc,'html')


        return Response({}, status = status.HTTP_200_OK)


from django.core.mail import send_mail , EmailMessage
from django.core.mail import EmailMultiAlternatives
import sendgrid

from django.contrib.auth.tokens import PasswordResetTokenGenerator

token_generator = PasswordResetTokenGenerator()
# print "token : " , token_generator.make_token(User.objects.get(pk = 1))

# path = reverse("account_reset_password_from_key", kwargs=dict(uidb36=user_pk_to_url_str(user),                                      key=temp_key))


class BulkUserCreationAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self, request, format=None):
        print request.data,'aaaaaaaa'
        location = request.data["locationData"]
        url = location.split('/ERP')[0]
        wb = load_workbook(filename = BytesIO(request.FILES['xl'].read()))
        ws = wb.worksheets[0]
        row_count = ws.max_row+1
        column_count = ws.max_column

        for i in range(2, row_count):
            try:
                username = ws['A' + str(i)].value
            except:
                username =""
            try:
                email = ws['B' + str(i)].value
            except:
                email =""
            try:
                first_name = ws['C' + str(i)].value
            except:
                first_name =" "
            # print first_name,'aaaaa'
            try:
                last_name = ws['D' + str(i)].value
                if last_name == None:
                    last_name = first_name
                else:
                    last_name = last_name
            except:
                last_name = first_name

            try:
                mobile = ws['E' + str(i)].value
            except:
                mobile =""
            try:
                designation = ws['F' + str(i)].value
                if designation == 'manager' or 'admin' or 'director':
                    is_staff = True
                else:
                    is_staff = False
            except:
                designation =""
            try:
                send = User(username=username, email= email, first_name=first_name, last_name=last_name,is_staff=is_staff)
                send.save()
            except:
                continue
            pobj = profile.objects.get(pk=send.profile.pk)
            pobj.email = email
            pobj.mobile = mobile
            pobj.details = {"username":username,"email":email,"first_name":first_name,"last_name":last_name,"designation":designation,"mobile":mobile,"GST":""}
            pobj.save()
            ctx = {
                'heading' : "Welcome to " + globalSettings.SITE_ADDRESS ,
                'link' : url + '/accounts/password/reset/',
                'recieverName' : first_name + ' ' + last_name,
                'brandName' : globalSettings.BRAND_NAME,
                'siteAddress' : globalSettings.SITE_ADDRESS
            }
            sendAddr = []
            sendAddr.append(str(email))
            email_body = get_template('app.user.resetPassword.html').render(ctx)
            email_subject = "Welcome to " + globalSettings.SITE_ADDRESS
            email_cc = []
            email_bcc = []
            send_email(email_body,sendAddr,email_subject,email_cc,email_bcc,'html')

        return Response(status = status.HTTP_200_OK)

@csrf_exempt
def socialMobileView(request):
    print request.POST,'ddddddddddddddddd'
    if request.POST['secretKey'] == globalSettings.MOBILE_SECRET_KEY:
        u = User.objects.filter(username = request.POST['username'])
        if len(u)>0:
            print 'already exist'
        else:
            print 'create new'
            u = User(username = request.POST['username'])
            u.email = request.POST['email']
            fname = request.POST['email'].split('@')[0]
            u.first_name = fname
            u.is_active = True
            u.set_password(request.POST['password'])
            u.save()
        loginView(request)
    else:
        raise PermissionDenied()
    if globalSettings.VERSION_ID == 1:
        page = "ngEcommerce.html"
    else:
        page = "ngEcommerceNext.html"
    return render(request,page)
